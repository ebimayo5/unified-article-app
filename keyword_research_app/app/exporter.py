from __future__ import annotations

import csv
import io
from pathlib import Path

from app.models import DOKUSOU_LIKE_COLUMNS, KeywordInput, KeywordResult, OUTPUT_COLUMNS


KEYWORD_INPUT_COLUMNS = ["keyword", "monthly_search_volume", "genre", "memo", "priority"]
KEYWORD_ALIASES = {
    "keyword",
    "keywords",
    "query",
    "search query",
    "キーワード",
    "検索語句",
}
VOLUME_ALIASES = {
    "volume",
    "search volume",
    "monthly volume",
    "monthly_search_volume",
    "月間検索数",
    "検索ボリューム",
    "ボリューム",
}
POSITION_ALIASES = {"position", "pos.", "pos", "rank", "ranking", "順位"}
URL_ALIASES = {"url", "page url", "traffic page", "流入URL", "ページURL"}
KD_ALIASES = {"kd", "keyword difficulty", "difficulty", "難易度"}

TEMPLATE_ROWS = [
    ["フォレスターXT 速すぎ", 390, "車", "中古購入前の不安系", "high"],
    ["セコム ホームセキュリティ 値段", 720, "家づくり", "費用比較系", "medium"],
    ["ランドリーチェスト カビない", 110, "生活", "物販につながりやすい", "high"],
]


def read_keywords(path: Path) -> list[KeywordInput]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _read_csv(path)
    if suffix == ".xlsx":
        return _read_xlsx(path)
    raise ValueError("Input file must be .csv or .xlsx")


def write_keyword_inputs(keywords: list[KeywordInput], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=KEYWORD_INPUT_COLUMNS)
        writer.writeheader()
        for keyword in keywords:
            writer.writerow(
                {
                    "keyword": keyword.keyword,
                    "monthly_search_volume": keyword.monthly_search_volume or "",
                    "genre": keyword.genre,
                    "memo": keyword.memo,
                    "priority": keyword.priority,
                }
            )


def _read_csv(path: Path) -> list[KeywordInput]:
    text = _read_csv_text(path)
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;")
    except csv.Error:
        dialect = csv.excel
    rows = list(csv.DictReader(io.StringIO(text), dialect=dialect))
    for row in rows:
        row["__source_file"] = str(path)
    return _rows_to_keywords(rows)


def _read_csv_text(path: Path) -> str:
    data = path.read_bytes()
    if data.startswith((b"\xff\xfe", b"\xfe\xff")):
        return data.decode("utf-16")
    if data.startswith(b"\xef\xbb\xbf"):
        return data.decode("utf-8-sig")
    for encoding in ("utf-8-sig", "cp932", "utf-16"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def _read_xlsx(path: Path) -> list[KeywordInput]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("XLSX input requires openpyxl. Run: pip install openpyxl") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    values = list(sheet.iter_rows(values_only=True))
    if not values:
        return []

    headers = [str(value).strip() if value is not None else "" for value in values[0]]
    rows = []
    for row in values[1:]:
        item = {}
        for index, header in enumerate(headers):
            item[header] = "" if index >= len(row) or row[index] is None else row[index]
        item["__source_file"] = str(path)
        rows.append(item)
    return _rows_to_keywords(rows)


def _rows_to_keywords(rows: list[dict]) -> list[KeywordInput]:
    column_map = _detect_columns(rows[0] if rows else {})
    keyword_column = column_map.get("keyword")
    if rows and not keyword_column:
        raise ValueError("Input file must contain a keyword column")

    keywords = []
    for row in rows:
        keyword = str(row.get(keyword_column, "")).strip() if keyword_column else ""
        if not keyword:
            continue
        keywords.append(
            KeywordInput(
                keyword=keyword,
                genre=str(_value(row, column_map.get("genre")) or "").strip(),
                memo=str(_value(row, column_map.get("memo")) or "").strip(),
                priority=str(_value(row, column_map.get("priority")) or "").strip(),
                monthly_search_volume=_parse_int(_value(row, column_map.get("volume"))),
                competitor_position=_parse_int(_value(row, column_map.get("position"))),
                competitor_url=str(_value(row, column_map.get("url")) or "").strip(),
                keyword_difficulty=_parse_float(_value(row, column_map.get("kd"))),
                source_file=str(row.get("__source_file", "") or "").strip(),
            )
        )
    return keywords


def _detect_columns(row: dict) -> dict[str, str]:
    normalized = {_normalize_header(key): key for key in row.keys()}
    return {
        "keyword": _find_column(normalized, KEYWORD_ALIASES),
        "volume": _find_column(normalized, VOLUME_ALIASES),
        "position": _find_column(normalized, POSITION_ALIASES),
        "url": _find_column(normalized, URL_ALIASES),
        "kd": _find_column(normalized, KD_ALIASES),
        "genre": _find_column(normalized, {"genre", "ジャンル"}),
        "memo": _find_column(normalized, {"memo", "メモ", "notes", "note"}),
        "priority": _find_column(normalized, {"priority", "優先度"}),
    }


def _find_column(normalized: dict[str, str], aliases: set[str]) -> str:
    for alias in aliases:
        key = _normalize_header(alias)
        if key in normalized:
            return normalized[key]
    for normalized_key, original in normalized.items():
        if any(_normalize_header(alias) in normalized_key for alias in aliases):
            return original
    return ""


def _normalize_header(value: object) -> str:
    return str(value or "").replace("\ufeff", "").strip().lower().replace("_", " ").replace("-", " ")


def _value(row: dict, column: str | None) -> object:
    if not column:
        return ""
    return row.get(column, "")


def _parse_int(value: object) -> int:
    if value is None:
        return 0
    text = str(value).strip().replace(",", "")
    if not text or text in {"-", "—"}:
        return 0
    try:
        return int(float(text))
    except ValueError:
        return 0


def _parse_float(value: object) -> float:
    if value is None:
        return 0.0
    text = str(value).strip().replace(",", "").replace("%", "")
    if not text or text in {"-", "—"}:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def write_csv(results: list[KeywordResult], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=OUTPUT_COLUMNS)
        writer.writeheader()
        for result in results:
            writer.writerow(result.to_row())


def write_xlsx(results: list[KeywordResult], path: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError("XLSX output requires openpyxl. Run: pip install openpyxl") from exc

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "results"
    sheet.append(DOKUSOU_LIKE_COLUMNS)
    sheet.freeze_panes = "A2"

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(name="Meiryo UI", size=10, color="FFFFFF", bold=True)
    sheet.row_dimensions[1].height = 24
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for result in results:
        row = result.to_dokusou_like_row()
        if result.opportunity_level == "未調査":
            row[DOKUSOU_LIKE_COLUMNS[17]] = ""
            row[DOKUSOU_LIKE_COLUMNS[19]] = ""
            row[DOKUSOU_LIKE_COLUMNS[20]] = ""
        sheet.append([row[column] for column in DOKUSOU_LIKE_COLUMNS])

    aim_col = DOKUSOU_LIKE_COLUMNS.index("ねらい目判定") + 1
    score_col = DOKUSOU_LIKE_COLUMNS.index("スコア") + 1
    level_col = DOKUSOU_LIKE_COLUMNS.index("判定レベル") + 1
    title_cols = {
        DOKUSOU_LIKE_COLUMNS.index("キーワード") + 1,
        DOKUSOU_LIKE_COLUMNS.index("月間検索数") + 1,
        aim_col,
        score_col,
        level_col,
        DOKUSOU_LIKE_COLUMNS.index("allintitle_count") + 1,
        DOKUSOU_LIKE_COLUMNS.index("intitle_count") + 1,
    }
    fills = {
        "best": PatternFill("solid", fgColor="D9EAD3"),
        "good": PatternFill("solid", fgColor="E2F0D9"),
        "watch": PatternFill("solid", fgColor="FFF2CC"),
        "hard": PatternFill("solid", fgColor="FCE4D6"),
        "unresearched": PatternFill("solid", fgColor="E5E7EB"),
        "qa": PatternFill("solid", fgColor="00B050"),
        "free": PatternFill("solid", fgColor="0070C0"),
        "tiktok": PatternFill("solid", fgColor="7030A0"),
        "instagram": PatternFill("solid", fgColor="C00000"),
        "x": PatternFill("solid", fgColor="7F7F7F"),
        "threads": PatternFill("solid", fgColor="BF9000"),
        "facebook": PatternFill("solid", fgColor="1F4E79"),
    }
    thin_border = Border(
        left=Side(style="thin", color="111827"),
        right=Side(style="thin", color="111827"),
        top=Side(style="thin", color="111827"),
        bottom=Side(style="thin", color="111827"),
    )
    site_pairs = {
        "qa": (3, 4),
        "free": (5, 6),
        "tiktok": (7, 8),
        "instagram": (9, 10),
        "x": (11, 12),
        "threads": (13, 14),
        "facebook": (15, 16),
    }
    centered_cols = set(range(2, 22))
    text_left_cols = {1, 22}
    for row_index in range(2, sheet.max_row + 1):
        sheet.row_dimensions[row_index].height = 22
        aim_value = str(sheet.cell(row=row_index, column=aim_col).value or "")
        score_value = sheet.cell(row=row_index, column=score_col).value or 0
        level_value = str(sheet.cell(row=row_index, column=level_col).value or "")

        for column_index in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row_index, column=column_index)
            cell.border = thin_border
            cell.font = Font(name="Meiryo UI", size=10, color="111827", bold=True)
            if column_index in text_left_cols:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=column_index == 22)
            elif column_index in centered_cols:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(vertical="center")

        for fill_key, columns in site_pairs.items():
            marker_cell = sheet.cell(row=row_index, column=columns[0])
            rank_cell = sheet.cell(row=row_index, column=columns[1])
            if marker_cell.value or rank_cell.value:
                for column_index in columns:
                    cell = sheet.cell(row=row_index, column=column_index)
                    cell.fill = fills[fill_key]
                    cell.font = Font(name="Meiryo UI", size=10, color="FFFFFF", bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center")

        if "\u672a\u8abf\u67fb" in aim_value or "\u672a\u8abf\u67fb" in level_value:
            for column_index in (aim_col, score_col, level_col):
                sheet.cell(row=row_index, column=column_index).fill = fills["unresearched"]
        elif "\u304b\u306a\u308a" in aim_value or score_value >= 80:
            sheet.cell(row=row_index, column=aim_col).fill = fills["best"]
        elif "\u72d9\u3044\u76ee" in aim_value or score_value >= 60:
            sheet.cell(row=row_index, column=aim_col).fill = fills["good"]
        elif "\u8981\u691c\u8a0e" in aim_value or score_value >= 40:
            sheet.cell(row=row_index, column=aim_col).fill = fills["watch"]
        elif score_value or level_value:
            sheet.cell(row=row_index, column=aim_col).fill = fills["hard"]

        for column_index in (aim_col, score_col, level_col):
            sheet.cell(row=row_index, column=column_index).alignment = Alignment(horizontal="center", vertical="center")
            sheet.cell(row=row_index, column=column_index).font = Font(name="Meiryo UI", size=10, color="111827", bold=True)

    widths = {
        "キーワード": 34,
        "月間検索数": 14,
        "Q&Aサイト": 12,
        "Q&A順位": 10,
        "無料ブログ": 12,
        "無料ブログ順位": 14,
        "ねらい目判定": 14,
        "スコア": 10,
        "判定レベル": 14,
        "allintitle_count": 18,
        "intitle_count": 16,
        "メモ": 52,
        "ジャンル": 16,
        "Q&A件数": 12,
        "無料ブログ件数": 16,
        "SNS件数": 12,
        "強いサイト件数": 16,
        "弱いサイト件数": 16,
        "検索URL": 48,
        "上位URL": 80,
    }
    for column_index, column_name in enumerate(DOKUSOU_LIKE_COLUMNS, start=1):
        width = widths.get(column_name, 13)
        sheet.column_dimensions[get_column_letter(column_index)].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    workbook.save(path)


def create_keyword_template(path: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError as exc:
        raise RuntimeError("Template creation requires openpyxl. Run: pip install openpyxl") from exc

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "keywords"
    sheet.append(KEYWORD_INPUT_COLUMNS)
    for row in TEMPLATE_ROWS:
        sheet.append(row)

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    widths = {
        "keyword": 36,
        "monthly_search_volume": 18,
        "genre": 16,
        "memo": 34,
        "priority": 14,
    }
    for index, column in enumerate(KEYWORD_INPUT_COLUMNS, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = widths[column]

    priority_validation = DataValidation(
        type="list",
        formula1='"high,medium,low"',
        allow_blank=True,
    )
    sheet.add_data_validation(priority_validation)
    priority_validation.add("E2:E500")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = "A1:E500"

    guide = workbook.create_sheet("使い方")
    guide["A1"] = "Keyword Treasure Finder 入力テンプレート"
    guide["A1"].font = Font(size=15, bold=True)
    guide["A3"] = "1. keyword列に調査したいキーワードを入力します。"
    guide["A4"] = "2. monthly_search_volume列にはAhrefsなどの月間検索数を入れられます。空でも実行できます。"
    guide["A5"] = "3. genre、memo、priorityは任意です。"
    guide["A6"] = "4. アプリはキーワードを通常検索し、月間検索数も結果に反映します。"
    guide.column_dimensions["A"].width = 92

    workbook.save(path)


def ensure_keyword_template(path: Path) -> None:
    if not path.exists():
        create_keyword_template(path)
        return

    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("Template update requires openpyxl. Run: pip install openpyxl") from exc

    workbook = load_workbook(path)
    sheet = workbook.active
    headers = [str(cell.value or "").strip() for cell in sheet[1]]
    if "monthly_search_volume" in headers:
        workbook.close()
        return
    if "keyword" in headers:
        keyword_index = headers.index("keyword") + 1
        insert_at = keyword_index + 1
    else:
        insert_at = 2
    sheet.insert_cols(insert_at)
    sheet.cell(row=1, column=insert_at).value = "monthly_search_volume"
    sheet.column_dimensions[sheet.cell(row=1, column=insert_at).column_letter].width = 18
    workbook.save(path)
